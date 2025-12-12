from app.services import GenerationService
import io
import openpyxl

# Mock Config
class MockConfig:
    CUSTOMERS_FILE = 'data/customers.json'
    TEMPLATE_FILE = 'data/template.csv'
    ADDRESSES_SHEET_URL = None

import app.services
app.services.Config = MockConfig

def verify_excel_text_format():
    service = GenerationService()
    
    # 1. VERIFY ORDERS
    params = {
        "cantidad_ordenes": 3,
        "pais": "Chile",
        "ciudad": "Santiago",
        "ct_origen": "CD TEST",
        "capacidad_min": 0.0010,
        "capacidad_max": 0.0050
    }
    
    # Needs valid customers
    service.customers = [
        {"address": "Test St", "country": "Chile", "city": "Santiago", "lat": -33.1, "long": -70.1, "id": "123"}
    ]
    
    print("Generating Orders Excel...")
    mem, filename = service.generate_excel(params)
    
    # Load with openpyxl to check format
    wb = openpyxl.load_workbook(mem)
    ws = wb.active
    
    # Check Header
    if ws['A1'].value == "N° DOCUMENTO":
        print("✅ Orders Header Match")
    else:
        print(f"❌ Orders Header Mismatch: {ws['A1'].value}")
        
    # Check Data Format
    cell_cap = ws['O2'] # Capacidad Uno
    print(f"DEBUG: Ord Row 2 Capacidad Value: '{cell_cap.value}' | Format: '{cell_cap.number_format}'")
    
    if cell_cap.number_format == '@':
        print("✅ Orders Cell Format is TEXT (@)")
    else:
        print(f"❌ Orders Cell Format Invalid: {cell_cap.number_format}")

    if ',' in str(cell_cap.value):
        print("✅ Orders Decimal Comma Verified")
    else:
        print(f"❌ Orders Decimal Comma Missing: {cell_cap.value}")


    # 2. VERIFY VEHICLES
    groups = [{"type": "Moto", "count": 1, "capacity1": 50}]
    print("\nGenerating Vehicles Excel...")
    mem_v, filename_v = service.generate_vehicles_excel(groups)
    
    wb_v = openpyxl.load_workbook(mem_v)
    ws_v = wb_v.active
    
    if ws_v['A1'].value == "PLACA":
        print("✅ Vehicles Header Match")
        
    cell_v_cap = ws_v['D2']
    print(f"DEBUG: Vec Row 2 Cap 1 Value: '{cell_v_cap.value}' | Format: '{cell_v_cap.number_format}'")
    
    if cell_v_cap.number_format == '@':
        print("✅ Vehicles Cell Format is TEXT (@)")
    else:
        print(f"❌ Vehicles Cell Format Invalid: {cell_v_cap.number_format}")

if __name__ == "__main__":
    verify_excel_text_format()
