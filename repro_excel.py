import sys
import os

# Mock Config environment before import
os.environ['CUSTOMERS_FILE'] = 'dummy.json' # Won't use
os.environ['TEMPLATE_FILE'] = 'data/plantilla_ordenes.csv'

# Setup paths (hacky for looking like package)
sys.path.append(os.getcwd())

from app.services import GenerationService
from app.config import Config

# Mock current_app
class MockLogger:
    def info(self, msg): print(f"[INFO] {msg}")
    def error(self, msg): print(f"[ERROR] {msg}")
    def warning(self, msg): print(f"[WARN] {msg}")

class MockApp:
    logger = MockLogger()
    config = {'LOG_LEVEL': 'INFO'}

import flask
flask.current_app = MockApp()

def test_generation():
    print("--- Starting Test ---")
    service = GenerationService()
    
    # 1. Test Header Loading
    print("Loading headers...")
    service._load_template_headers()
    print(f"Headers detected: {len(service.headers)}")
    print(f"First 3 headers: {service.headers[:3]}")
    
    if len(service.headers) < 2:
        print("❌ FAIL: Headers broken (likely delimiter issue)")
        sys.exit(1)
    
    # 2. Mock Customers (skip network)
    service.customers = [
        {"address": "Calle Falsa 123", "country": "Chile", "city": "Santiago", "lat": "-33", "long": "-70", "id": "TEST-1"}
    ]
    
    # 3. Generate
    print("Generating Excel...")
    try:
        f, name = service.generate_excel({
            "cantidad_ordenes": 5, 
            "ct_origen": "TEST-CD",
            "pais": "Chile",
            "ciudad": "Santiago"
        })
        print("✅ Success! File generated.")
        
        # Verify content? (Hard without saving, but no crash is good start)
        # We could use openpyxl to read 'f' back
        from openpyxl import load_workbook
        wb = load_workbook(f)
        ws = wb.active
        print(f"Rows in sheet: {ws.max_row}")
        
        if ws.max_row > 1:
            print("✅ Data rows found!")
        else:
            print("❌ FAIL: Only header row found.")
            
    except Exception as e:
        print(f"❌ Crash: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_generation()
