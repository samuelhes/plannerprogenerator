import pytest
import os
from app import create_app
from app.services import GenerationService
from openpyxl import load_workbook
import io

@pytest.fixture
def client():
    app = create_app()
    app.config['TESTING'] = True
    with app.test_client() as client:
        yield client

def test_api_generation_success_custom_geo(client):
    payload = {
        "cantidad_ordenes": 5,
        "items_por_orden": 1,
        "ct_origen": "TEST_CT",
        "capacidad_min": 0.001,
        "capacidad_max": 0.005,
        "ventana_inicio": "08:00",
        "ventana_fin": "12:00",
        "city": "Ciudad Inventada",
        "country": "Pais Nuevo"
    }
    response = client.post('/api/generate', json=payload)
    assert response.status_code == 200

def test_multi_row_items():
    service = GenerationService()
    params = {
        "cantidad_ordenes": 2,
        "items_por_orden": 3, 
        "ct_origen": "TEST_CT",
        "capacidad_min": 1,
        "capacidad_max": 2
    }
    output, _ = service.generate_excel(params)
    wb = load_workbook(filename=output)
    ws = wb.active
    assert ws.max_row == 7 

def test_api_generation_missing_ct_origen(client):
    payload = {
        "cantidad_ordenes": 1
    }
    response = client.post('/api/generate', json=payload)
    assert response.status_code == 400

def test_capacity2_generation():
    service = GenerationService()
    params = {
        "cantidad_ordenes": 5,
        "ct_origen": "TEST",
        "capacidad_min": 0.001,
        "capacidad_max": 0.002,
        # Cap 2 Enabled
        "capacidad2_min": 0.010,
        "capacidad2_max": 0.020
    }
    output, _ = service.generate_excel(params)
    wb = load_workbook(filename=output)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    idx_cap2 = headers.index("CAPACIDAD DOS")
    
    for row in ws.iter_rows(min_row=2):
        cap2 = row[idx_cap2].value
        assert cap2 is not None
        assert "," in cap2
        val = float(cap2.replace(',', '.'))
        assert 0.010 <= val <= 0.020

def test_dynamic_tags_generation():
    service = GenerationService()
    params = {
        "cantidad_ordenes": 10,
        "ct_origen": "TEST",
        "capacidad_min": 1,
        "capacidad_max": 2,
        "tags": [
            {"header": "TIPO_CAMION", "values": ["A", "B"]},
            {"header": "PRIORIDAD", "values": ["Alta", "Baja"]}
        ]
    }
    output, _ = service.generate_excel(params)
    wb = load_workbook(filename=output)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    # Check headers exist
    assert "TIPO_CAMION" in headers
    assert "PRIORIDAD" in headers
    
    idx_tipo = headers.index("TIPO_CAMION")
    
    for row in ws.iter_rows(min_row=2):
        val = row[idx_tipo].value
        assert val in ["A", "B"]
