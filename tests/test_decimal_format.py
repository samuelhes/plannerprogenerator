
import pytest
from app.services import GenerationService
import openpyxl
import io

class TestDecimalFormat:
    def setup_method(self):
        self.service = GenerationService()
        # Mock customers to avoid external calls
        self.service.customers = [{
            "address": "Calle Falsa 123",
            "country": "Chile",
            "city": "Santiago",
            "lat": "-33.4489",
            "long": "-70.6693",
            "name": "Test Client",
            "id": "123"
        }]

    def test_generate_excel_decimal_format(self):
        """
        Verify that generated Order Excel uses dots for decimals.
        """
        params = {
            "cantidad_ordenes": 5,
            "ct_origen": "CD Test",
            "capacidad_min": 1.5,
            "capacidad_max": 2.5,
            "capacidad2_min": 0.5,
            "capacidad2_max": 0.9,
            "ventana_inicio": "09:00",
            "ventana_fin": "18:00"
        }
        
        output, filename = self.service.generate_excel(params)
        wb = openpyxl.load_workbook(output)
        ws = wb.active
        
        # Get Header Map
        headers = [cell.value for cell in ws[1]]
        idx_cap1 = headers.index("CAPACIDAD UNO")
        idx_cap2 = headers.index("CAPACIDAD DOS")
        idx_lat = headers.index("LATITUD")
        
        # Check rows
        for row in ws.iter_rows(min_row=2):
            cap1 = row[idx_cap1].value
            cap2 = row[idx_cap2].value
            lat = row[idx_lat].value
            
            # Assertions
            assert "," not in str(cap1), f"Capacity 1 contains comma: {cap1}"
            assert "." in str(cap1), f"Capacity 1 should have dot: {cap1}"
            
            assert "," not in str(cap2), f"Capacity 2 contains comma: {cap2}"
            assert "." in str(cap2), f"Capacity 2 should have dot: {cap2}"
            
            assert "," not in str(lat), f"Latitude contains comma: {lat}"
            # Latitude might be empty if not in mocked data, but here it is.

    def test_generate_vehicles_decimal_format(self):
        """
        Verify that generated Vehicle Excel uses dots for decimals.
        """
        groups = [{
            "type": "Camion",
            "count": 2,
            "capacity1": 1000.5,
            "capacity2": 500.25,
            "origin": "CD Test",
            "start_time": "08:00",
            "end_time": "18:00"
        }]
        
        output, filename = self.service.generate_vehicles_excel(groups)
        wb = openpyxl.load_workbook(output)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        idx_cap1 = headers.index("CAPACIDAD UNO")
        idx_cap2 = headers.index("CAPACIDAD DOS")
        idx_period = headers.index("PERIODO DE RECARGA [HORAS]")
        
        for row in ws.iter_rows(min_row=2):
            cap1 = row[idx_cap1].value
            cap2 = row[idx_cap2].value
            period = row[idx_period].value
            
            assert "," not in str(cap1), f"Vehicle Cap1 contains comma: {cap1}"
            assert "." in str(cap1), f"Vehicle Cap1 should have dot: {cap1}"
            
            assert "," not in str(cap2), f"Vehicle Cap2 contains comma: {cap2}"
            assert "." in str(cap2), f"Vehicle Cap2 should have dot: {cap2}"
            
            assert "," not in str(period), f"Period contains comma: {period}"
            assert "." in str(period), f"Period should have dot: {period}"
            assert str(period) == "0.25", f"Period should be 0.25, got {period}"
