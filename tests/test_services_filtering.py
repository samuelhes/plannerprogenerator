from unittest.mock import MagicMock, patch
from app.services import GenerationService
from openpyxl import load_workbook

def test_sheet_filtering_logic():
    # Mock data to simulate Sheet content
    mock_data = [
        {"address": "Calle 1", "country": "Chile", "city": "Santiago"},
        {"address": "Calle 2", "country": "Chile", "city": "Valparaíso"},
        {"address": "Calle 3", "country": "Argentina", "city": "Mendoza"}
    ]
    
    service = GenerationService()
    service.customers = mock_data # Inject mock data directly handling logic
    
    # helper for filtering
    def get_filtered(p, c):
        # We need to replicate the filter logic or just call generate?
        # Let's call generate_excel with 1 order and check the result or 
        # better yet, expose filter logic? 
        # Actually, let's verify via generate_excel result counting if possible 
        # OR just copy the filter logic check here since it's inside the method
        # But `service.generate_excel` creates a file. 
        # Let's inspect `filtered_customers` by patching random.choice?
        pass

    # Scenario 1: Specific City
    # Params: Country=Chile, City=Santiago
    # Expected: Only Calle 1
    params_santiago = {
        "cantidad_ordenes": 10, 
        "pais": "Chile", 
        "ciudad": "Santiago",
        "ct_origen": "TEST"
    }
    
    # We can detect what was filtered by checking the output addresses
    output1, _ = service.generate_excel(params_santiago)
    wb1 = load_workbook(output1)
    ws1 = wb1.active
    # Check that all addresses are Calle 1 or contain it?
    # Our data has 'Calle 1'.
    
    for row in ws1.iter_rows(min_row=2):
        addr = row[12].value # DIRECCION column index might vary, let's search header
        # Actually simplest: mock random.choice to see what list it got?
        pass

# A better unit test for the service
def test_filtering_strict():
    service = GenerationService()
    service.customers = [
        {"address": "SCL1", "country": "Chile", "city": "Santiago"},
        {"address": "VAL1", "country": "Chile", "city": "Valparaiso"},
        {"address": "ARG1", "country": "Argentina", "city": "Mendoza"}
    ]
    
    # Test 1: Chile + Santiago
    params = {"pais": "Chile", "ciudad": "Santiago", "cantidad_ordenes": 1, "ct_origen": "T", "capacidad_min":1, "capacidad_max":2}
    # We can't access local var 'filtered_customers' easily. 
    # But we can assert the generated file matches.
    out, _ = service.generate_excel(params)
    wb = load_workbook(out)
    ws = wb.active
    # Address is usually col index 3 (0-based) ?? No.
    # Let's find header index
    headers = [c.value for c in ws[1]]
    idx_addr = headers.index("DIRECCION")
    val = ws.cell(row=2, column=idx_addr+1).value
    assert val == "SCL1"

def test_filtering_wildcard_country():
    service = GenerationService()
    service.customers = [
        {"address": "SCL1", "country": "Chile", "city": "Santiago"},
        {"address": "VAL1", "country": "Chile", "city": "Valparaiso"},
        {"address": "ARG1", "country": "Argentina", "city": "Mendoza"}
    ]
    
    # Test: Chile + "Otro" (or no city) -> Should get SCL1 OR VAL1
    params = {"pais": "Chile", "ciudad": "Otro", "cantidad_ordenes": 20, "ct_origen": "T", "capacidad_min":1, "capacidad_max":2}
    out, _ = service.generate_excel(params)
    wb = load_workbook(out)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx_addr = headers.index("DIRECCION")
    
    found_scl = False
    found_val = False
    found_arg = False
    
    for row in ws.iter_rows(min_row=2):
        val = row[idx_addr].value
        if val == "SCL1": found_scl = True
        if val == "VAL1": found_val = True
        if val == "ARG1": found_arg = True
        
    assert found_scl or found_val
    assert not found_arg # Should NOT contain Argentina
