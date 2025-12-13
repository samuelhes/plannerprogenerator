"""
Unit tests for services module
"""
import pytest
import json
from io import BytesIO
from app.services import GenerationService
from openpyxl import load_workbook


@pytest.fixture
def service():
    """Create a GenerationService instance"""
    return GenerationService()


@pytest.fixture
def valid_order_params():
    """Valid order generation parameters"""
    return {
        'cantidad_ordenes': 10,
        'ct_origen': 'CD Central',
        'fecha_entrega': '2025-01-15',
        'capacidad_min': 1.0,
        'capacidad_max': 10.0,
        'ventana_inicio': '09:00',
        'ventana_fin': '18:00',
        'pais': 'Chile',
        'ciudad': 'Santiago'
    }


@pytest.fixture
def valid_vehicle_groups():
    """Valid vehicle groups"""
    return [
        {
            'count': 5,
            'type': 'Camion',
            'origin': 'CD Norte',
            'capacity1': 1000,
            'capacity2': None,
            'start_time': '08:00',
            'end_time': '18:00'
        }
    ]


class TestGenerateExcel:
    """Tests for generate_excel method"""
    
    def test_generate_excel_success(self, service, valid_order_params):
        """Test successful Excel generation"""
        output, filename = service.generate_excel(valid_order_params)
        
        assert output is not None
        assert isinstance(output, BytesIO)
        assert filename.endswith('.xlsx')
        assert 'ordenes' in filename.lower()
    
    def test_generate_excel_column_order(self, service, valid_order_params):
        """Test that columns are in correct order"""
        output, _ = service.generate_excel(valid_order_params)
        
        # Load the generated Excel
        wb = load_workbook(output)
        ws = wb.active
        
        expected_headers = [
            "N° DOCUMENTO", "LATITUD", "LONGITUD", "DIRECCION", "NOMBRE ITEM", 
            "CANTIDAD", "CODIGO ITEM", "FECHA MIN ENTREGA", "FECHA MAX ENTREGA",
            "MIN VENTANA HORARIA 1", "MAX VENTANA HORARIA 1", 
            "MIN VENTANA HORARIA 2", "MAX VENTANA HORARIA 2",
            "COSTO ITEM", "CAPACIDAD UNO", "CAPACIDAD DOS", "SERVICE TIME",
            "IMPORTANCIA", "IDENTIFICADOR CONTACTO", "NOMBRE CONTACTO",
            "TELEFONO", "EMAIL CONTACTO", "CT ORIGEN"
        ]
        
        # Get actual headers from first row
        actual_headers = [cell.value for cell in ws[1]]
        
        assert actual_headers == expected_headers, f"Header mismatch: {actual_headers}"
    
    def test_generate_excel_text_format(self, service, valid_order_params):
        """Test that all cells are formatted as text"""
        output, _ = service.generate_excel(valid_order_params)
        
        wb = load_workbook(output)
        ws = wb.active
        
        # Check a few cells for text format (@)
        for row in ws.iter_rows(min_row=2, max_row=3):
            for cell in row:
                if cell.value is not None:
                    assert cell.number_format == '@', f"Cell {cell.coordinate} not text format"
    
    def test_generate_excel_missing_ct_origen(self, service, valid_order_params):
        """Test that missing CT Origen raises error"""
        params = valid_order_params.copy()
        params['ct_origen'] = ''
        
        with pytest.raises(ValueError, match="CT Origen"):
            service.generate_excel(params)
    
    def test_generate_excel_row_count(self, service, valid_order_params):
        """Test correct number of rows generated"""
        params = valid_order_params.copy()
        params['cantidad_ordenes'] = 5
        params['items_por_orden'] = 2
        
        output, _ = service.generate_excel(params)
        
        wb = load_workbook(output)
        ws = wb.active
        
        # Header + (5 orders * 2 items) = 11 rows
        assert ws.max_row == 11


class TestGenerateVehiclesExcel:
    """Tests for generate_vehicles_excel method"""
    
    def test_generate_vehicles_success(self, service, valid_vehicle_groups):
        """Test successful vehicle Excel generation"""
        output, filename = service.generate_vehicles_excel(valid_vehicle_groups)
        
        assert output is not None
        assert isinstance(output, BytesIO)
        assert filename == 'flota_vehiculos.xlsx'
    
    def test_generate_vehicles_column_count(self, service, valid_vehicle_groups):
        """Test correct number of columns"""
        output, _ = service.generate_vehicles_excel(valid_vehicle_groups)
        
        wb = load_workbook(output)
        ws = wb.active
        
        # Should have 21 columns + empty trailing
        assert ws.max_column >= 21
    
    def test_generate_vehicles_row_count(self, service, valid_vehicle_groups):
        """Test correct number of vehicles generated"""
        groups = valid_vehicle_groups.copy()
        groups[0]['count'] = 3
        
        output, _ = service.generate_vehicles_excel(groups)
        
        wb = load_workbook(output)
        ws = wb.active
        
        # Header + 3 vehicles = 4 rows
        assert ws.max_row == 4
    
    def test_generate_vehicles_text_format(self, service, valid_vehicle_groups):
        """Test that all cells are formatted as text"""
        output, _ = service.generate_vehicles_excel(valid_vehicle_groups)
        
        wb = load_workbook(output)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, max_row=2):
            for cell in row:
                if cell.value is not None:
                    assert cell.number_format == '@'


class TestLocalization:
    """Tests for localization features"""
    
    def test_get_localized_name_latam(self, service):
        """Test LATAM name generation"""
        name = service._get_localized_name('Chile')
        
        assert isinstance(name, str)
        assert len(name.split()) == 2  # First + Last name
    
    def test_get_localized_name_us(self, service):
        """Test US name generation"""
        name = service._get_localized_name('United States')
        
        assert isinstance(name, str)
        assert len(name.split()) == 2
