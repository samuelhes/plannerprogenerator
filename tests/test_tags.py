
import pytest
from app.services import GenerationService
import openpyxl
import io

class TestDynamicTags:
    def setup_method(self):
        self.service = GenerationService()
        self.service.customers = [{
            "address": "Calle Falsa 123",
            "country": "Chile",
            "city": "Santiago",
            "lat": "-33.4489",
            "long": "-70.6693",
            "name": "Test Client",
            "id": "123"
        }]

    def test_generate_excel_with_tags(self):
        params = {
            "cantidad_ordenes": 5,
            "ct_origen": "CD Test",
            "capacidad_min": 1,
            "capacidad_max": 2,
            "fecha_entrega": "2023-01-01",
            "tags": [
                {"header": "TIPO_CLIENTE", "values": ["VIP", "Regular"]},
                {"header": "ZONA", "values": ["Norte", "Sur"]}
            ]
        }
        
        output, filename = self.service.generate_excel(params)
        wb = openpyxl.load_workbook(output)
        ws = wb.active
        
        headers = [c.value for c in ws[1]]
        
        # Verify Headers
        assert "TIPO_CLIENTE" in headers
        assert "ZONA" in headers
        
        idx_type = headers.index("TIPO_CLIENTE")
        idx_zona = headers.index("ZONA")
        
        # Verify Values
        for row in ws.iter_rows(min_row=2):
            type_val = row[idx_type].value
            zona_val = row[idx_zona].value
            
            assert type_val in ["VIP", "Regular"]
            assert zona_val in ["Norte", "Sur"]

    def test_generate_vehicles_with_tags(self):
        payload = {
            "groups": [{
                "type": "Camion",
                "count": 2,
                "capacity1": 1000,
                "origin": "CD",
                "start_time":"08:00",
                "end_time":"18:00"
            }],
            "tags": [
                 {"header": "PROVEEDOR", "values": ["TransA", "TransB"]}
            ]
        }
        
        output, filename = self.service.generate_vehicles_excel(payload)
        wb = openpyxl.load_workbook(output)
        ws = wb.active
        
        headers = [c.value for c in ws[1]]
        assert "PROVEEDOR" in headers
        
        idx_prov = headers.index("PROVEEDOR")
        
        for row in ws.iter_rows(min_row=2):
            val = row[idx_prov].value
            assert val in ["TransA", "TransB"]
