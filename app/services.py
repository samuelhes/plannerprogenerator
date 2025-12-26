import os
import io
import csv
import json
import random
import requests
import itertools
from datetime import datetime
from openpyxl import Workbook
from .config import Config

from flask import current_app

class GenerationService:
    def __init__(self):
        # Load Resources on init
        # Note: In a real production app, we might want to lazy load or cache these better,
        # but for now, we'll log the init.
        pass

    def _get_logger(self):
        # specific helper to access current_app logger safely
        try:
            return current_app.logger
        except:
             # Fallback for scripts running outside context if ever needed
            import logging
            return logging.getLogger(__name__)

    def _load_customers(self):
        if hasattr(self, 'customers') and self.customers:
            return

        # 1. Try Google Sheet
        sheet_data = self._load_from_sheet()
        if sheet_data:
            self._get_logger().info("Loaded addresses from Google Sheet.")
            self.customers = sheet_data
            return

        # 2. Fallback to Local JSON
        try:
            with open(Config.CUSTOMERS_FILE, 'r', encoding='utf-8') as f:
                self.customers = json.load(f)
                self._get_logger().info("Loaded addresses from local JSON.")
        except Exception as e:
            self._get_logger().error(f"Error loading customers from JSON: {e}")
            self.customers = []

    def _load_template_headers(self):
        # Deprecated: Headers are now strictly enforced by code
        pass

    def _load_from_sheet(self):
        url = Config.ADDRESSES_SHEET_URL
        if not url:
            return None
        
        try:
            # Add timeout for production standards
            response = requests.get(url, timeout=10)
            response.raise_for_status()
            
            # Parse CSV
            # Data structure: Col A=Address, Col B=Country, Col C=City
            decoded_content = response.content.decode('utf-8')
            cr = csv.reader(decoded_content.splitlines(), delimiter=',')
            
            customers = []
            rows = list(cr)
            if not rows: return None
            
            for row in rows:
                if len(row) < 3: continue
                # Basic normalization
                addr = row[0].strip()
                country = row[1].strip()
                city = row[2].strip()
                
                # Simple heuristc to skip header
                if "direccion" in addr.lower() and "pais" in country.lower():
                    continue

                # Senior QA: Ensure critical data isn't empty
                if not addr or not country:
                     continue

                customers.append({
                    "address": addr,
                    "country": country,
                    "city": city,
                    "lat": "", # Sheet doesn't have it
                    "long": "",
                    "name": "Cliente Sheet", # Generic name
                    "id": f"S-{random.randint(1000,9999)}"
                })
            return customers

        except Exception as e:
            self._get_logger().error(f"Error fetching/parsing Google Sheet: {e}")
            return None

    def _get_localized_name(self, country_input):
        # Database of names
        LOCALE_DATA = {
            "LATAM": {
                "names": ["Juan", "Maria", "Carlos", "Ana", "Jose", "Luis", "Sofia", "Camila", "Pedro", "Diego"],
                "surnames": ["Gonzalez", "Rodriguez", "Perez", "Fernandez", "Lopez", "Diaz", "Martinez", "Silva", "Rojas", "Soto"]
            },
            "US": {
                "names": ["John", "Mary", "Michael", "Jennifer", "James", "Linda", "Robert", "Patricia", "David", "Elizabeth"],
                "surnames": ["Smith", "Johnson", "Williams", "Jones", "Brown", "Davis", "Miller", "Wilson", "Moore", "Taylor"]
            }
        }
        
        # Determine Region
        c = country_input.lower()
        if any(x in c for x in ["chile", "argentina", "colombia", "mexico", "peru", "bolivia", "uruguay", "ecuador", "venezuela", "paraguay"]):
            region = "LATAM"
        else:
            region = "US" # Default/Global
            
        data = LOCALE_DATA[region]
        return f"{random.choice(data['names'])} {random.choice(data['surnames'])}"

    def generate_excel(self, params: dict):
        # ... (Previous Resource Loading)
        self._load_customers()
        
        # HEADERS (Strict)
        HEADERS = [
            "N° DOCUMENTO", "LATITUD", "LONGITUD", "DIRECCION", "NOMBRE ITEM", "CANTIDAD", "CODIGO ITEM",
            "FECHA MIN ENTREGA", "FECHA MAX ENTREGA", "MIN VENTANA HORARIA 1", "MAX VENTANA HORARIA 1",
            "MIN VENTANA HORARIA 2", "MAX VENTANA HORARIA 2", "COSTO ITEM", "CAPACIDAD UNO", "CAPACIDAD DOS",
            "SERVICE TIME", "IMPORTANCIA", "IDENTIFICADOR CONTACTO", "NOMBRE CONTACTO", "TELEFONO", "EMAIL CONTACTO",
            "CT ORIGEN"
        ]
        
        # APPEND DYNAMIC HEADERS
        tags = params.get('tags', [])
        tag_headers = [t['header'] for t in tags]
        HEADERS.extend(tag_headers)
        
        # Finally append empty column for strictness if desired, or just end
        HEADERS.append("")

        # Destructure params (Keep existing logic)
        try:
            count = int(params.get('cantidad_ordenes', 40))
        except (ValueError, TypeError):
            count = 40
        try:
            items_per_order = int(params.get('items_por_orden') or 1)
        except (ValueError, TypeError):
            items_per_order = 1
        if items_per_order < 1: items_per_order = 1
        try:
            cap_min = float(params.get('capacidad_min', 1))
            cap_max = float(params.get('capacidad_max', 10))
        except (ValueError, TypeError):
            cap_min = 1.0
            cap_max = 10.0
        if cap_min > cap_max:
             cap_min, cap_max = cap_max, cap_min
        cap2_min = params.get('capacidad2_min')
        cap2_max = params.get('capacidad2_max')
        win1_start = params.get('ventana_inicio', '09:00')
        win1_end = params.get('ventana_fin', '18:00')
        win2_start = params.get('ventana2_inicio')
        win2_end = params.get('ventana2_fin')
        ct_origin = params.get('ct_origen')
        if not ct_origin or str(ct_origin).strip() == "":
             raise ValueError("CT Origen is mandatory.")
        service_time = params.get('service_time')
        city_filter = params.get('ciudad', '')
        country_filter = params.get('pais', '')
        delivery_date = params.get('fecha_entrega', datetime.now().strftime('%Y-%m-%d'))
        try:
            date_obj = datetime.strptime(delivery_date, '%Y-%m-%d')
            formatted_date = date_obj.strftime('%d/%m/%Y')
        except:
            formatted_date = delivery_date

        # Filter customers (Keep existing logic)
        filtered_customers = []
        target_country = country_filter.lower().strip()
        target_city = city_filter.lower().strip()
        for c in self.customers:
            c_country = c.get('country', '').lower().strip()
            c_city = c.get('city', '').lower().strip()
            if target_country and target_country not in c_country and "otro" not in target_country: continue
            if target_city and "otro" not in target_city:
                if target_city not in c_city: continue
            filtered_customers.append(c)
        if not filtered_customers: filtered_customers = self.customers
        if not filtered_customers: raise ValueError("No customer data available.")
        random.shuffle(filtered_customers)
        customer_iterator = itertools.cycle(filtered_customers)

        # WEBWORKBOOK GENERATION
        wb = Workbook()
        ws = wb.active
        ws.title = "Ordenes Generadas"
        
        ws.append(HEADERS)

        global_item_counter = 1

        for i in range(count):
            customer = next(customer_iterator)
            contact_name = self._get_localized_name(customer.get('country', country_filter))
            order_id = f"ORD-{str(i+1).zfill(6)}"
            cap_val = random.uniform(cap_min, cap_max)
            cap_str = f"{cap_val:.4f}"
            cap2_str = ""
            if cap2_min is not None and cap2_max is not None:
                c2_val = random.uniform(float(cap2_min), float(cap2_max))
                cap2_str = f"{c2_val:.4f}"

            for j in range(items_per_order):
                row_map = {}
                row_map["N° DOCUMENTO"] = order_id
                row_map["LATITUD"] = str(customer.get('lat', ''))
                row_map["LONGITUD"] = str(customer.get('long', ''))
                row_map["DIRECCION"] = customer.get('address', '')
                row_map["NOMBRE ITEM"] = f"Item {global_item_counter}"
                row_map["CANTIDAD"] = "1"
                row_map["CODIGO ITEM"] = f"SKU-{global_item_counter}"
                row_map["COSTO ITEM"] = "1500"
                global_item_counter += 1
                row_map["FECHA MIN ENTREGA"] = formatted_date
                row_map["FECHA MAX ENTREGA"] = formatted_date
                row_map["MIN VENTANA HORARIA 1"] = win1_start
                row_map["MAX VENTANA HORARIA 1"] = win1_end
                row_map["MIN VENTANA HORARIA 2"] = win2_start if win2_start else ""
                row_map["MAX VENTANA HORARIA 2"] = win2_end if win2_end else ""
                row_map["CAPACIDAD UNO"] = cap_str
                row_map["CAPACIDAD DOS"] = cap2_str
                row_map["SERVICE TIME"] = str(service_time) if service_time else "5"
                row_map["IMPORTANCIA"] = "1"
                row_map["IDENTIFICADOR CONTACTO"] = customer.get('id', '')
                row_map["NOMBRE CONTACTO"] = contact_name
                row_map["TELEFONO"] = f"569{random.randint(11111111,99999999)}"
                row_map["EMAIL CONTACTO"] = f"contacto{i}@example.com"
                row_map["CT ORIGEN"] = ct_origin
                row_map["CT ORIGEN"] = ct_origin
                
                # DYNAMIC TAG VALUES
                for t in tags:
                    header = t['header']
                    values = t['values']
                    if values:
                        row_map[header] = random.choice(values)
                    else:
                        row_map[header] = ""

                row_map[""] = "" 

                row_values = [row_map.get(h, "") for h in HEADERS]
                ws.append(row_values)

        # FORCE TEXT FORMAT
        for row in ws.iter_rows():
            for cell in row:
                cell.number_format = '@'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output, f"ordenes_{count}.xlsx"

    def generate_vehicles_excel(self, vehicle_groups):
        """
        Generates a fleet Excel file (.xlsx) with Text Format.
        """
        HEADERS = [
            "PLACA", "ORIGEN", "DESTINO", "CAPACIDAD UNO", "CAPACIDAD DOS", 
            "HORA INICIO JORNADA", "HORA FIN JORNADA", "INICIO HORA DESCANSO", 
            "FIN HORA DESCANSO", "COSTO POR SALIDA", "COSTO POR KILOMETRO", 
            "COSTO POR HORA", "COSTO FIJO", "MAXIMA CANTIDAD DE ENTREGAS POR RECORRIDO", 
            "MAXIMO TIEMPO DE MANEJO [HORAS]", "MAXIMA CANTIDAD DE RECORRIDOS", 
            "DISTANCIA MAXIMA POR RECORRIDO [KILOMETROS]", "VELOCIDAD VEHICULO", 
            "PERIODO DE RECARGA [HORAS]", "MAXIMO DE DINERO", "NO CONSIDERAR RETORNO AL CD",
            "PERIODO DE RECARGA [HORAS]", "MAXIMO DE DINERO", "NO CONSIDERAR RETORNO AL CD"
        ]
        
        # APPEND DYNAMIC HEADERS (VEHICLES)
        # Note: input is a list of groups. Should we take tags from the first group? 
        # Or should tags be a top-level param? 
        # The frontend usually sends structure: { groups: [...] }. 
        # We need to change the API expectations or assume tags come in specific way.
        # Let's assume the API will receive { groups: [...], tags: [...] } in the new iteration.
        # But 'generate_vehicles_excel' signature currently takes 'vehicle_groups' list.
        # We need to refactor the route calling this. 
        # For now, let's gracefully handle if vehicle_groups is a dict containing 'groups' and 'tags', 
        # or just a list (backward compat).
        
        vehicle_tags = []
        groups_list = []
        
        if isinstance(vehicle_groups, dict):
             groups_list = vehicle_groups.get('groups', [])
             vehicle_tags = vehicle_groups.get('tags', [])
        else:
             groups_list = vehicle_groups
             
        tag_headers = [t['header'] for t in vehicle_tags]
        HEADERS.extend(tag_headers)
        HEADERS.append("")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Flota Generada"
        ws.append(HEADERS)
        
        prefix_map = { "Moto": "MOTO", "Auto": "AUTO", "Camion": "CAMI", "Bici": "BICI", "Otro": "OTRO" }
        counters = {k: 0 for k in prefix_map.values()}
        
        for group in groups_list:
            v_type = group.get('type', 'Otro')
            count = int(group.get('count', 0))
            cap1 = group.get('capacity1', '')
            cap2 = group.get('capacity2', '')
            origin = group.get('origin', '')
            start_time = group.get('start_time', '')
            end_time = group.get('end_time', '')
            
            prefix = prefix_map.get(v_type, "OTRO")
            if v_type not in prefix_map: 
                prefix = v_type[:4].upper()
                if prefix not in counters: counters[prefix] = 0

            for _ in range(count):
                counters[prefix] += 1
                sequence = counters[prefix]
                plate = f"{prefix}{str(sequence).zfill(2)}"
                
                row_map = {
                    "PLACA": plate,
                    "ORIGEN": origin,
                    "DESTINO": "",
                    "CAPACIDAD UNO": str(cap1) if cap1 else "",
                    "CAPACIDAD DOS": str(cap2) if cap2 else "",
                    "HORA INICIO JORNADA": start_time,
                    "HORA FIN JORNADA": end_time,
                    "INICIO HORA DESCANSO": "",
                    "FIN HORA DESCANSO": "",
                    "COSTO POR SALIDA": "1000",
                    "COSTO POR KILOMETRO": "333",
                    "COSTO POR HORA": "11",
                    "COSTO FIJO": "666",
                    "MAXIMA CANTIDAD DE ENTREGAS POR RECORRIDO": "111",
                    "MAXIMO TIEMPO DE MANEJO [HORAS]": "20",
                    "MAXIMA CANTIDAD DE RECORRIDOS": "4",
                    "DISTANCIA MAXIMA POR RECORRIDO [KILOMETROS]": "500",
                    "VELOCIDAD VEHICULO": "Normal",
                    "PERIODO DE RECARGA [HORAS]": "0.25",
                    "MAXIMO DE DINERO": "5500000",
                    "NO CONSIDERAR RETORNO AL CD": "1",
                    "NO CONSIDERAR RETORNO AL CD": "1"
                }
                
                # DYNAMIC TAG VALUES (VEHICLES)
                for t in vehicle_tags:
                    header = t['header']
                    values = t['values']
                    if values:
                        row_map[header] = random.choice(values)
                    else:
                        row_map[header] = ""
                        
                row_map[""] = ""
                row_values = [row_map.get(h, "") for h in HEADERS]
                ws.append(row_values)

        # FORCE TEXT FORMAT
        for row in ws.iter_rows():
            for cell in row:
                cell.number_format = '@'


        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output, "flota_vehiculos.xlsx"
